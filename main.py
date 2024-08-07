import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side
from tkinter import *
from tkinter import filedialog, messagebox
import os.path

def detect_header_and_extract_data(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # Attempt to find the header row by looking for specific column names
    header_row = None
    for i, row in df.iterrows():
        if 'Source' in row.values:
            header_row = i
            break

    if header_row is None:
        raise ValueError("Header row with 'Location' not found")
    
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    
    # Ensure that all required columns are present
    required_columns = ['Source', '4000g', '3500g', '2600g', 'Hours']
    if not all(col in df.columns for col in required_columns):
        raise ValueError(f"One or more required columns are missing in the sheet {sheet_name}")

    return df

def extract_data_create_new_excel(file_path, sheet_name):
    try:
        df = detect_header_and_extract_data(file_path, sheet_name)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process the Excel file\n{e}")
        return

    df.rename(columns={'Location': 'Source'}, inplace=True)

    columns = ['Source', '4000g', '3500g', '2600g', 'Hours']
    if not all(col in df.columns for col in columns):
        messagebox.showerror("Error", f"Expected columns {columns} not found in the sheet {sheet_name}")
        return

    data_to_extract = df[columns]
    df['4000g'] = pd.to_numeric(df['4000g'], errors='coerce')
    df['3500g'] = pd.to_numeric(df['3500g'], errors='coerce')
    df['2600g'] = pd.to_numeric(df['2600g'], errors='coerce')
    df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce')

    grouped_df = data_to_extract.groupby('Source').sum().reset_index()

    wb = Workbook()
    ws = wb.active

    # Add the custom header and make A1 bold with font size 14
    ws['A1'] = 'Enva'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = ''

    # Define border style
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Append the DataFrame to the worksheet starting from the fourth row
    for r_idx, row in enumerate(dataframe_to_rows(grouped_df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border  # Apply border to each cell
            cell.font = Font(bold=True, size=14)  # Make each cell bold with font size 14

    # Apply borders and bold font with size 14 to header manually
    for cell in ws[3]:
        cell.font = Font(bold=True, size=14)
        cell.border = thin_border

    # Apply borders and bold font with size 14 to every cell manually
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = thin_border

    try:
        # Get the user's home directory
        home_dir = os.path.expanduser("~")

        # Construct the path to the desktop
        desktop_path = os.path.join(home_dir, "Downloads")

        # Construct the full path for the output file
        output_file_path = os.path.join(desktop_path, "Enva Monthly.xlsx")
        print(f"Saving file to: {output_file_path}")

        # Save the workbook
        wb.save(output_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save the Excel file\n{e}")

    messagebox.showinfo("Success", f"Data has been saved to {output_file_path} with the desired formatting.")

def import_file():
    file_path = filedialog.askopenfilename(title="Select a file", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if file_path:
        sheet_name = sheet_name_entry.get()
        if not sheet_name:
            messagebox.showwarning("Input Required", "Please enter the sheet name.")
        else:
            extract_data_create_new_excel(file_path, sheet_name)

# Create the main Tkinter window
root = Tk()
root.title("Import File Example")
root.geometry("400x300")
root.configure(bg="#2e3f4f")  # Dark background color

# Style configurations
label_font = ("Helvetica", 12, "bold")
entry_font = ("Helvetica", 12)
button_font = ("Helvetica", 12, "bold")

# Create a frame for the content with padding
frame = Frame(root, bg="#2e3f4f")
frame.pack(pady=30, padx=30, fill="both", expand=True)

# Create and style widgets
sheet_name_label = Label(frame, text="Sheet Name:", font=label_font, bg="#2e3f4f", fg="#ffffff")
sheet_name_label.pack(pady=5)

sheet_name_entry = Entry(frame, font=entry_font, bd=2, relief="groove", bg="#1e2f3f", fg="#ffffff", insertbackground="#ffffff")
sheet_name_entry.pack(pady=10)

import_button = Button(frame, text="Import File", font=button_font, bg="#4caf50", fg="#ffffff", padx=10, pady=5, bd=0, relief="ridge", highlightthickness=0, activebackground="#45a049", cursor="hand2", command=import_file)
import_button.pack(pady=20)

# Apply custom styles to the button to make it look cooler
def style_button(button):
    button.config(
        borderwidth=0,
        relief="flat",
        overrelief="flat",
        highlightthickness=0
    )

style_button(import_button)

# Run the Tkinter event loop
root.mainloop()